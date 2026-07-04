# -*- coding: utf-8 -*-
"""
DiaspoInvest - Mise a jour mensuelle du Tracker Dashboard
=========================================================
Usage : python scripts/maj_tracker_mensuel.py

Ce script :
  1. Lit brvm_latest.json (depuis le depot automation local ou telecharge depuis GitHub)
  2. Met a jour la feuille "Cours Live" du Tracker Excel (colonne D = cours, colonne E = dividende brut)
  3. Sauvegarde une version datee : DiaspoInvest_Tracker_Dashboard_AAAA-MM.xlsx
  4. Affiche un recapitulatif des cours mis a jour

A lancer une fois par mois avant d'envoyer le fichier aux clients.
"""

import json
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erreur : installe openpyxl d'abord -> pip install openpyxl")
    sys.exit(1)

# --- Configuration -----------------------------------------------------------

TRACKER_SOURCE = Path(
    r"C:\Users\DJIOKAP JORDAN\DiaspoInvest\produits\VALIDES - PRETS A VENDRE"
    r"\DiaspoInvest Tracker Dashboard.xlsx"
)

# JSON BRVM : d'abord le depot local, sinon le depot automation
JSON_CANDIDATES = [
    Path(r"C:\Users\DJIOKAP JORDAN\diaspoinvest-automation\scripts\data\brvm_latest.json"),
    Path(r"C:\Users\DJIOKAP JORDAN\Downloads\brvm_latest.json"),
]

OUTPUT_DIR = Path(r"C:\Users\DJIOKAP JORDAN\DiaspoInvest\produits\VERSIONS CLIENTS")

# -----------------------------------------------------------------------------


def charger_json_brvm():
    for p in JSON_CANDIDATES:
        if p.exists():
            print(f"JSON BRVM : {p}")
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    print("ERREUR : brvm_latest.json introuvable. Verifie le depot automation.")
    sys.exit(1)


def construire_index_brvm(data):
    """Retourne un dict {code_action: {cours, dividende_brut}} depuis le JSON."""
    index = {}
    for action in data.get("actions", []):
        code = action.get("code", "").strip().upper()
        cours = action.get("cours")
        div = action.get("dividende_brut") or action.get("dividende") or 0
        if code and cours:
            index[code] = {"cours": cours, "dividende_brut": div}
    return index


def maj_cours_live(ws_form, index_brvm):
    """
    Met a jour la feuille Cours Live.
    Structure attendue : ligne 1 = en-tete, lignes 4+ = donnees actions
    Col B = code action, Col C = nom, Col D = cours (a mettre a jour), Col E = dividende brut
    """
    mises_a_jour = []
    non_trouves = []

    for row in range(4, ws_form.max_row + 1):
        code_cell = ws_form.cell(row, 2).value  # colonne B
        if not code_cell:
            continue
        code = str(code_cell).strip().upper()
        if code in index_brvm:
            ancien_cours = ws_form.cell(row, 4).value
            nouveau_cours = index_brvm[code]["cours"]
            div = index_brvm[code]["dividende_brut"]
            ws_form.cell(row, 4).value = nouveau_cours
            if div:
                ws_form.cell(row, 5).value = div
            mises_a_jour.append(
                f"  {code:<12} cours {ancien_cours or '?':>8} -> {nouveau_cours:>8}  |  div brut {div}"
            )
        else:
            nom = ws_form.cell(row, 3).value or ""
            non_trouves.append(f"  {code:<12} ({nom})")

    return mises_a_jour, non_trouves


def maj_date_dans_feuille(ws):
    """Met a jour la cellule de date de derniere MAJ dans Cours Live."""
    today_str = date.today().strftime("%d/%m/%Y")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and "derniere mise a jour" in v.lower():
                # Remplace la date dans la chaine
                parts = v.split("|")
                new_parts = []
                for p in parts:
                    if "mise a jour" in p.lower() or "mise à jour" in p.lower():
                        new_parts.append(f" Derniere mise a jour : {today_str} -- source : sikafinance.com  ")
                    else:
                        new_parts.append(p)
                cell.value = "|".join(new_parts)
                return


def main():
    print("=" * 60)
    print("DiaspoInvest - Mise a jour Tracker mensuel")
    print("=" * 60)

    # 1. Charger les donnees BRVM
    data = charger_json_brvm()
    date_brvm = data.get("genere_le", "inconnue")
    print(f"Donnees BRVM du : {date_brvm}")
    print(f"Nombre d'actions dans le JSON : {len(data.get('actions', []))}")

    index = construire_index_brvm(data)

    # 2. Charger le Tracker
    if not TRACKER_SOURCE.exists():
        print(f"ERREUR : Tracker introuvable -> {TRACKER_SOURCE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(TRACKER_SOURCE)

    # Trouver la feuille Cours Live
    nom_cours = None
    for name in wb.sheetnames:
        if "cours" in name.lower():
            nom_cours = name
            break

    if not nom_cours:
        print("ERREUR : feuille 'Cours Live' introuvable dans le Tracker")
        sys.exit(1)

    ws = wb[nom_cours]
    print(f"\nFeuille trouvee : '{nom_cours}' ({ws.max_row} lignes)")

    # 3. Mettre a jour
    mises_a_jour, non_trouves = maj_cours_live(ws, index)
    maj_date_dans_feuille(ws)

    print(f"\n{len(mises_a_jour)} actions mises a jour :")
    for ligne in mises_a_jour:
        print(ligne)

    if non_trouves:
        print(f"\n{len(non_trouves)} actions dans le Tracker mais absentes du JSON :")
        for ligne in non_trouves:
            print(ligne)

    # 4. Sauvegarder la version datee
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    mois = date.today().strftime("%Y-%m")
    output_path = OUTPUT_DIR / f"DiaspoInvest_Tracker_Dashboard_{mois}.xlsx"
    wb.save(output_path)

    print(f"\nFichier sauvegarde : {output_path}")
    print(f"Taille : {output_path.stat().st_size // 1024} KB")
    print("\nA envoyer aux clients via Brevo ou email direct.")
    print("=" * 60)


if __name__ == "__main__":
    main()
