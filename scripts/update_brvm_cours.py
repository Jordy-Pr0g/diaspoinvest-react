# coding: utf-8
"""
DiaspoInvest -- Mise a jour automatique des cours BRVM
Source : sikafinance.com/marches/aaz
Usage  : double-clic ou python update_brvm_cours.py
"""

import urllib.request
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
import sys
import re

FICHIER_EXCEL = r"C:\Users\DJIOKAP JORDAN\Desktop\DiaspoInvest - PERSONNEL\[PERSO] DiaspoInvest Tracker - Suivi de Portefeuille Boursier.xlsx"
URL_SIKA = "https://www.sikafinance.com/marches/aaz"

MAPPING = {
    "AFRICA GLOBAL LOGISTICS": "SDSC",
    "BANK OF AFRICA BENIN": "BOAB",
    "BANK OF AFRICA BURKINA FASO": "BOABF",
    "BANK OF AFRICA CI": "BOAC",
    "BANK OF AFRICA MALI": "BOAM",
    "BANK OF AFRICA NIGER": "BOAN",
    "BANK OF AFRICA SENEGAL": "BOAS",
    "BANQUE INTERNATIONALE POUR LE COMMERCE DU BENIN": "BICB",
    "BERNABE": "BNBC",
    "BICICI": "BICC",
    "CFAO CI": "CFAC",
    "CIE CI": "CIEC",
    "CORIS BANK INTERNATIONAL BF": "CBIBF",
    "CROWN SIEM": "SEMC",
    "ECOBANK CI": "ECOC",
    "ERIUM": "SIVC",
    "ETI TG": "ETIT",
    "FILTISAC CI": "FTSC",
    "LOTERIE NATIONALE DU BENIN": "LNBB",
    "MOVIS CI": "MVSC",
    "NEI CEDA CI": "NEIC",
    "NESTLE CI": "NTLC",
    "NSIA BANQUE": "NSBC",
    "ONATEL BF": "ONTBF",
    "ORANGE CI": "ORAC",
    "ORAGROUP TOGO": "ORGT",
    "PALMCI": "PALC",
    "SAFCA CI": "SAFC",
    "SAPH CI": "SPHC",
    "SERVAIR ABIDJAN CI": "ABJC",
    "SETAO CI": "STAC",
    "SGBCI": "SGBC",
    "SICABLE CI": "CABC",
    "SICOR": "SICC",
    "SITAB": "STBC",
    "SMB CI": "SMBC",
    "SOCIETE IVOIRIENNE DE BANQUE CI": "SIBC",
    "SODECI": "SDCC",
    "SOGB": "SOGB",
    "SOLIBRA CI": "SLBC",
    "SONATEL": "SNTS",
    "SUCRIVOIRE": "SCRC",
    "TOTAL CI": "TTLC",
    "TOTAL SENEGAL": "TTLS",
    "TRACTAFRIC MOTORS CI": "PRSC",
    "UNILEVER CI": "UNLC",
    "UNIWAX CI": "UNXC",
    "VIVO ENERGY CI": "SVOC",
}


def normalise(text):
    # Remplace tous types d'espaces (nbsp, narrow-nbsp, etc.) par espace simple
    texte = re.sub(r"[\s   ​]+", " ", text)
    return texte.strip().upper()


def scrape_cours():
    print("Connexion a sikafinance.com...")
    req = urllib.request.Request(
        URL_SIKA,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if len(tables) < 2:
        raise ValueError("Structure sikafinance modifiee -- tableau actions introuvable.")

    table = tables[1]
    rows = table.find_all("tr")[1:]

    cours = {}
    non_mappes = []

    for row in rows:
        tds = row.find_all(["td", "th"])
        if len(tds) < 7:
            continue

        nom_sika = normalise(tds[0].get_text(separator=" ", strip=True))
        dernier_str = re.sub(r"[^0-9.,]", "", tds[6].get_text(strip=True)).replace(",", ".")

        try:
            prix = float(dernier_str)
        except ValueError:
            continue

        code = MAPPING.get(nom_sika)
        if code:
            cours[code] = prix
        else:
            non_mappes.append("  - %r => %s" % (nom_sika, prix))

    if non_mappes:
        print("[INFO] %d actions non mappees :" % len(non_mappes))
        for m in non_mappes:
            print(m)

    print("[OK] %d cours recuperes depuis sikafinance." % len(cours))
    return cours


def mettre_a_jour_excel(cours):
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb["Cours Live"]

    mises_a_jour = 0
    inchanges = 0
    non_trouves = []

    for row_idx in range(4, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=1).value
        if not code:
            continue

        if code in cours:
            ancien = ws.cell(row=row_idx, column=4).value
            nouveau = cours[code]
            ws.cell(row=row_idx, column=4).value = nouveau
            if ancien != nouveau:
                mises_a_jour += 1
                variation = ""
                if isinstance(ancien, (int, float)) and ancien > 0:
                    pct = (nouveau - ancien) / ancien * 100
                    variation = "  (%+.1f%%)" % pct
                ancien_str = str(int(ancien)) if isinstance(ancien, (int, float)) else str(ancien)
                print("  %-6s  %8s -> %8d%s" % (code, ancien_str, int(nouveau), variation))
            else:
                inchanges += 1
        else:
            non_trouves.append(code)

    ws["A2"] = "Derniere MAJ auto : %s -- sikafinance.com" % datetime.now().strftime("%d/%m/%Y %H:%M")

    wb.save(FICHIER_EXCEL)
    print("\n[OK] Fichier sauvegarde.")
    print("     %d cours mis a jour, %d inchanges." % (mises_a_jour, inchanges))
    if non_trouves:
        print("     %d codes sans correspondance : %s" % (len(non_trouves), non_trouves))


def main():
    print("=" * 55)
    print("  DiaspoInvest -- MAJ cours BRVM")
    print("  " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    print("=" * 55)

    try:
        cours = scrape_cours()
        mettre_a_jour_excel(cours)
        print("\nMise a jour terminee avec succes.")
    except urllib.error.URLError as e:
        print("\n[ERREUR] Impossible de joindre sikafinance : %s" % e)
        sys.exit(1)
    except FileNotFoundError:
        print("\n[ERREUR] Fichier Excel introuvable :\n  %s" % FICHIER_EXCEL)
        sys.exit(1)
    except Exception as e:
        print("\n[ERREUR] %s" % e)
        import traceback
        traceback.print_exc()
        sys.exit(1)

    input("\nAppuyez sur Entree pour fermer...")


if __name__ == "__main__":
    main()
